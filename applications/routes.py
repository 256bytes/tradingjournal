from flask import render_template, request, flash
from flask_login import current_user
from openpyxl import load_workbook

from applications import app
from applications.models import Brokers, Transactions
from applications import stock
from applications.database import db
from applications.brokerage_taxes import call_discount_brokers_buy, call_no_discount_brokers_buy, call_upstox_broker_buy


# @app.route('/my-link', methods=["GET","POST"])
# def import_data_form():

#     f = request.files['File']
#     f.save(f.filename)
#     fname = f.filename
#     book = load_workbook(fname)

#     sheet = book.active
#     for row in range(2, 2 + 1):
#         tc = sheet.cell(row, 1).value
#         type = sheet.cell(row, 2).value
#         script = sheet.cell(row, 3).value
#         price = sheet.cell(row, 4).value
#         qty = sheet.cell(row, 5).value

#         chk_tc = db.session.query(Brokers.trading_code).filter(Brokers.trading_code == tc).first()
#         broker_type = db.session.query(Brokers.type).filter(Brokers.trading_code == tc).first()
#         broker_name = db.session.query(Brokers.name).filter(Brokers.trading_code == tc).first()
#         if not chk_tc:
#             print("not found")
#         chk_script = stock.lookup(script)
#         print(f"this is chk_script: {chk_script}")
        
#         if not chk_script:
#             print("invalid symbol")
#         if type == "CNC":
#             if broker_type:
#                 result = call_discount_brokers_buy(price, qty, tc)
#             if broker_type == False and broker_name.name != "UPSTOX":
#                 result = call_no_discount_brokers_buy(price, qty, tc)
#             if broker_name.name == "UPSTOX":
#                 result = call_upstox_broker_buy(price, qty, tc)

#             script_to_add = Transactions(user_id = current_user.id,
#                         type = "CNC",
#                         call = 'Buy',
#                         script = script,
#                         price = price,
#                         qty = qty,
#                         brokerage_per_unit = result['brokerage'],
#                         net_rate_per_unit = result['unit_per_after_brokerage'],
#                         net_total_before_levies = result['net_total_after_brokerage'],
#                         transaction_chgs = result['tran_chgs'],
#                         dp_chgs = result['dp_chgs'],
#                         stt = result['s_trans'],
#                         sebi_turnover_fees = result['sebi_chgs'],
#                         stamp_duty = result['sduty'],
#                         gst = result['gs_tax'],
#                         total_taxes = result['tot_tax'],
#                         net_total = result['pay'],
#                         broker = broker_name.name, # square braket to extract the data from the tupple
#                         trading_code = tc
#                         )
#             db.session.add(script_to_add)
#             db.session.commit()
#             flash("Script successfully added!", category='success')
#             return "sucess"
#         else:
#             pass

#         print('\n\n')
#         print(current_user.id)
#         print(chk_tc)
#         print(broker_type)
#         print(broker_name)
#         print(type)
#         print(script)
#         print(price)
#         print(qty)
#         print('\n\n')
    

#     return "sucess"
# @app.route('/test')
# def test_page():
#     return render_template('test.html')

