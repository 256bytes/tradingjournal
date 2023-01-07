from flask import request, flash, redirect, url_for
from flask_login import current_user, login_required
from openpyxl import load_workbook
import os

#-------------User Packages --------------------#
from applications import app
from applications.models import Analysts, Research
from applications.helpers import lookup
from applications.database import db

@app.route('/analysis_import', methods=["GET","POST"])
@login_required
def import_analysis_form():

    if not request.files['File']:
        flash("Select the file to import", category='warning')
        return redirect(url_for('analytics_page'))
    try:
        f = request.files['File']
        f.save(f.filename)
        fname = f.filename
        book = load_workbook(fname)

        sheet = book.active

    except Exception as e:
        flash(f"Something went wrong while importing the file. error: {e}", category='warning')
        return redirect(url_for('analytics_page'))

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value:

            script = sheet.cell(row, 1).value
            price = sheet.cell(row, 2).value
            call = sheet.cell(row, 3).value
            sl = sheet.cell(row, 4).value
            tgt = sheet.cell(row, 5).value
            tmf = sheet.cell(row, 6).value
            analyst = sheet.cell(row, 7).value
            resource = sheet.cell(row, 8).value
            chk_script = lookup(script)
            call = call.upper()
            
            if not chk_script:
                flash(f'Invalid {script} symbol', category="warning")
                os.remove(f.filename)
                return redirect(url_for('analytics_page'))
                break

            get_analyst = Analysts.query.filter(Analysts.name == analyst).first()

            if not (get_analyst):
                add_analyst = Analysts(name = analyst,
                                    number_of_calls = 1
                                    )
                db.session.add(add_analyst)
                db.session.commit()
            else:
                get_calls = db.session.query(Analysts.number_of_calls).filter(Analysts.name == analyst).first()
                get_analyst.number_of_calls = get_calls.number_of_calls +1

                db.session.add(get_analyst)
                db.session.commit()

            db.session.rollback()
            db.session.begin()
            add_research = Research (user_id = current_user.id,
                                    script = script,
                                    price = price,
                                    call = call,
                                    stop_loss = sl,
                                    target = tgt,
                                    time_frame = tmf,
                                    analyst = analyst,
                                    resource = resource)
            db.session.add(add_research)
            db.session.commit()
        else:
            os.remove(f.filename)
            break
        
    os.remove(f.filename)

    flash("Successfully added data to research table", category="success")  
    return redirect(url_for('analytics_page'))
