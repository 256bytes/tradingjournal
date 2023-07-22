from flask import request, flash, redirect, url_for
from flask_login import current_user, login_required
from openpyxl import load_workbook
import os

from applications import app


@app.route("/data_import", methods=["GET", "POST"])
@login_required
def import_data_form():
    # -------------User Packages --------------------#
    from applications.calc_taxes.remove_taxes import RemoveBrokerage

    call_type = "Buy"

    if not request.files["File"]:
        flash("Select the file to import", category="warning")
        return redirect(url_for("holdings_page", page=1))

    try:
        from applications.helpers import SymbolLookup, chk_special

        f = request.files["File"]
        f.save(f.filename)
        fname = f.filename
        book = load_workbook(fname)

        sheet = book.active

    except Exception as e:
        flash(f"Something went wrong while importing the file. error: {e}", category="warning")
        os.remove(f.filename)
        return redirect(url_for("holdings_page", page=1))

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value:
            try:
                from applications.validate_database import ValidateUserDatabase

                user_broker_data = ValidateUserDatabase()

                trading_code = sheet.cell(row, 1).value
                result, message = user_broker_data.validate_trader_code(current_user.id, trading_code)
                print()
                print(result)
                print(message)
                print()
                if not result:
                    flash(message, category="danger")
                    os.remove(f.filename)
                    return redirect(url_for("home_page"))
                else:
                    trade_mode = sheet.cell(row, 2).value
                    symbol = sheet.cell(row, 3).value
                    symbol = symbol.upper()
                    from applications.helpers import SymbolLookup

                    validate_stock_symbol = SymbolLookup().find_symbol(symbol)
                    if validate_stock_symbol:
                        from applications.calc_taxes.remove_taxes import RemoveBrokerage

                        price = sheet.cell(row, 4).value
                        qty = sheet.cell(row, 5).value

                        price_without_brokerage = RemoveBrokerage(current_user.id, trading_code, price, qty)
                        price_without_brokerage = (
                            price_without_brokerage.r_rpu
                        )  # .r_rpu this is the variable instance of class RemoveBrokerage, which gives rate per uint.
                        from applications.user_database import GetUserData

                        data = GetUserData(
                            route="/data_import",
                            user_id=current_user.id,
                            trading_code=trading_code,
                            symbol=symbol,
                            call_type=call_type,
                            price=price_without_brokerage,
                            qty=qty,
                            trade_mode=trade_mode,
                        )
                        result, message = data.brokerage_and_taxes()
                        if not result:
                            flash(message, category="danger")
                            os.remove(f.filename)
                            return redirect(url_for("holdings_page", page=1))
                        else:
                            result, message = data.add_transactions()
                            if not result:
                                flash(message, category="danger")
                                os.remove(f.filename)
                                return redirect(url_for("holdings_page", page=1))
                            else:
                                continue
                    else:
                        flash(message, category="danger")
                        os.remove(f.filename)
                        return redirect(url_for("home_page"))
            except Exception as e:
                flash(f"Something went wrong while reading the file. error: {e}", category="warning")
                os.remove(f.filename)
                return redirect(url_for("holdings_page", page=1))
        else:
            os.remove(f.filename)
            break
    flash("Script successfully added!", category="success")
    return redirect(url_for("holdings_page", page=1))
